#!/usr/bin/env python3
"""
Importeer vlogs uit daphnevlogs_overzicht_v6.xlsx naar de website.
Haalt automatisch de duration op via de YouTube Data API v3.

Stappen:
1. Kopieer scripts/config.example.json naar scripts/config.json
2. Voeg je YouTube API key toe in scripts/config.json
3. Plaats Excel in ~/Downloads/daphnevlogs_overzicht_v6.xlsx
4. Run: python3 scripts/import_excel.py
"""
import json
import os
import re
import ssl
import sys
import urllib.error
import urllib.request
from urllib.parse import parse_qs, urlparse

try:
    import certifi
    SSL_CONTEXT = ssl.create_default_context(cafile=certifi.where())
except ImportError:
    SSL_CONTEXT = None

try:
    import openpyxl
except ImportError:
    print("Installeer openpyxl: pip3 install openpyxl")
    sys.exit(1)

CONFIG_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'config.json')

COUNTRY_MAP = {
    'Paris': ('fr', 'Frankrijk'),
    'Haarlem': ('nl', 'Nederland'),
    'Ibiza': ('es', 'Spanje'),
    'Dubrovnik': ('hr', 'Kroatië'),
    'America': ('us', 'Verenigde Staten'),
    'Indonesia': ('id', 'Indonesië'),
    'Austria': ('at', 'Oostenrijk'),
    'Camping Bulgaria': ('bg', 'Bulgarije'),
    'Romania': ('ro', 'Roemenië'),
    'Valduggia': ('it', 'Italië'),
    'London': ('gb', 'Verenigd Koninkrijk'),
    'Cambridge': ('gb', 'Verenigd Koninkrijk'),
    'Oxford': ('gb', 'Verenigd Koninkrijk'),
    'Rome': ('it', 'Italië'),
    'Antwerpen': ('be', 'België'),
    'Budapest': ('hu', 'Hongarije'),
    'Gent': ('be', 'België'),
    'IJsselmeer': ('nl', 'Nederland'),
    'Linz': ('at', 'Oostenrijk'),
    'Breda': ('nl', 'Nederland'),
    'Amsterdam': ('nl', 'Nederland'),
    'New Orleans': ('us', 'Verenigde Staten'),
    'Bulgaria': ('bg', 'Bulgarije'),
    'Italy': ('it', 'Italië'),
    'Utrecht': ('nl', 'Nederland'),
    'Volendam': ('nl', 'Nederland'),
    'Berlin': ('de', 'Duitsland'),
    'Lloret de Mar': ('es', 'Spanje'),
    'Singapore': ('sg', 'Singapore'),
    'Madurodam': ('nl', 'Nederland'),
}

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(SCRIPT_DIR)
EXCEL_PATH = os.path.expanduser('~/Downloads/daphnevlogs_overzicht_v6.xlsx')


def get_country(title, sheet_name):
    if sheet_name == 'Mixed_Media':
        return ('nl', 'Nederland')
    title_lower = (title or '').lower()
    for key, (code, name) in COUNTRY_MAP.items():
        if key.lower() in title_lower:
            return (code, name)
    return ('nl', 'Nederland')


def make_id(title, year, sheet_name, idx):
    safe = re.sub(r'[^a-z0-9]+', '-', (title or 'vlog').lower())
    safe = safe.strip('-')[:25]
    prefix = 'mixed' if sheet_name == 'Mixed_Media' else 'vacation'
    return f"{prefix}-{year}-{safe}-{idx}"


def esc_js(s):
    if s is None:
        return ''
    return str(s).replace('\\', '\\\\').replace("'", "\\'").replace('\n', ' ').replace('\r', '')


def load_api_key():
    """Laad YouTube API key uit scripts/config.json"""
    example_path = os.path.join(os.path.dirname(CONFIG_PATH), 'config.example.json')
    if not os.path.exists(CONFIG_PATH) and os.path.exists(example_path):
        import shutil
        shutil.copy(example_path, CONFIG_PATH)
        print("Eerste run: scripts/config.json aangemaakt. Voeg je YouTube API key toe en run opnieuw.")
        return None
    if not os.path.exists(CONFIG_PATH):
        return None
    try:
        with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
            cfg = json.load(f)
        key = (cfg.get('youtube_api_key') or '').strip()
        return key if key and key != 'VOEG_HIER_JOUW_YOUTUBE_API_KEY_TOE' else None
    except (json.JSONDecodeError, OSError):
        return None


def youtube_video_id(url):
    """Haal YouTube video ID uit URL (youtu.be of youtube.com)."""
    if not url or not isinstance(url, str):
        return None
    url = url.strip()
    try:
        if 'youtu.be/' in url:
            return url.split('youtu.be/')[-1].split('?')[0].split('/')[0] or None
        if 'youtube.com' in url:
            parsed = urlparse(url)
            return parse_qs(parsed.query).get('v', [None])[0]
    except Exception:
        pass
    return None


def parse_iso8601_duration(iso_str):
    """Parse YouTube ISO 8601 duration (PT1H2M10S) naar mm:ss of hh:mm:ss."""
    if not iso_str or not isinstance(iso_str, str):
        return None
    # PT1H2M10S, PT10M30S, PT30S, etc.
    m = re.match(r'PT(?:(\d+)H)?(?:(\d+)M)?(?:(\d+)S)?', iso_str.upper())
    if not m:
        return None
    h = int(m.group(1) or 0)
    m_min = int(m.group(2) or 0)
    s = int(m.group(3) or 0)
    if h > 0:
        return f"{h}:{m_min:02d}:{s:02d}"
    return f"{m_min}:{s:02d}" if (m_min or s) else None


def fetch_youtube_durations(vlogs, api_key):
    """
    Haal durations op via YouTube Data API v3.
    Retourneert dict: {video_id: 'mm:ss' of 'hh:mm:ss'}
    """
    if not api_key:
        return {}
    video_ids = []
    url_to_id = {}
    for v in vlogs:
        vid = youtube_video_id(v.get('url'))
        if vid:
            video_ids.append(vid)
            url_to_id[vid] = vid
    if not video_ids:
        return {}

    result = {}
    # YouTube API: max 50 ids per request
    batch_size = 50
    for i in range(0, len(video_ids), batch_size):
        batch = video_ids[i:i + batch_size]
        ids_param = ','.join(batch)
        url = f"https://www.googleapis.com/youtube/v3/videos?part=contentDetails&id={ids_param}&key={api_key}"
        try:
            req = urllib.request.Request(url)
            kwargs = {'timeout': 15}
            if SSL_CONTEXT:
                kwargs['context'] = SSL_CONTEXT
            with urllib.request.urlopen(req, **kwargs) as resp:
                data = json.loads(resp.read().decode())
            for item in data.get('items', []):
                vid = item.get('id')
                dur = item.get('contentDetails', {}).get('duration')
                if vid and dur:
                    parsed = parse_iso8601_duration(dur)
                    if parsed:
                        result[vid] = parsed
        except (urllib.error.HTTPError, urllib.error.URLError, json.JSONDecodeError, OSError) as e:
            print(f"Waarschuwing: YouTube API fout ({e}), durations blijven '—'")
            break
    return result


def main():
    if not os.path.exists(EXCEL_PATH):
        print(f"Excel-bestand niet gevonden: {EXCEL_PATH}")
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    all_vlogs = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(h).lower().replace(' ', '_') if h else f"col_{i}" for i, h in enumerate(rows[0])]

        for idx, row in enumerate(rows[1:], 1):
            if not any(row):
                continue
            d = dict(zip(headers, row))
            year = d.get('year')
            title = d.get('title')
            url = d.get('youtube_link')
            date_raw = d.get('date_raw')

            if not title:
                continue

            year = int(year) if year is not None and str(year).replace('.', '').isdigit() else 2024
            code, name = get_country(title, sheet_name)

            entry = {
                'id': make_id(title, year, sheet_name, idx),
                'title': title,
                'countryCode': code,
                'countryName': name,
                'dateRange': str(date_raw) if date_raw else str(year),
                'duration': '—',
                'url': url or '',
                'year': year,
                'isPopular': False,
                'isFavorite': False,
            }
            if sheet_name == 'Mixed_Media':
                entry['categories'] = ['event']
            if d.get('description'):
                entry['description'] = d['description']

            all_vlogs.append(entry)

    wb.close()
    all_vlogs.sort(key=lambda x: (-x['year'], x['title'] or ''))

    # Haal durations op via YouTube API
    api_key = load_api_key()
    if api_key:
        durations = fetch_youtube_durations(all_vlogs, api_key)
        for v in all_vlogs:
            vid = youtube_video_id(v.get('url'))
            if vid and vid in durations:
                v['duration'] = durations[vid]
        print(f"Durations opgehaald voor {len(durations)} video's")
    else:
        print("Geen YouTube API key gevonden in scripts/config.json – durations blijven '—'")
        print("Kopieer config.example.json naar config.json en voeg je API key toe.")

    # Genereer JavaScript VLOGS array
    lines = ['// VLOGS uit daphnevlogs_overzicht_v6.xlsx', 'const VLOGS = [']
    for v in all_vlogs:
        parts = [
            f"  {{",
            f"    id: '{esc_js(v['id'])}',",
            f"    title: '{esc_js(v['title'])}',",
            f"    countryCode: '{esc_js(v['countryCode'])}',",
            f"    countryName: '{esc_js(v['countryName'])}',",
            f"    dateRange: '{esc_js(v['dateRange'])}',",
            f"    duration: '{esc_js(v['duration'])}',",
            f"    url: '{esc_js(v['url'])}',",
            f"    year: {v['year']},",
            f"    isPopular: {str(v['isPopular']).lower()},",
            f"    isFavorite: {str(v['isFavorite']).lower()},",
        ]
        if v.get('categories'):
            parts.append(f"    categories: {json.dumps(v['categories'])},")
        if v.get('description'):
            parts.append(f"    description: '{esc_js(v['description'])}',")
        parts[-1] = parts[-1].rstrip(',')
        parts.append("  },")
        lines.append('\n'.join(parts))
    lines.append('];')

    vlogs_block = '\n'.join(lines)

    # Vervang VLOGS array in main.js
    main_js_path = os.path.join(PROJECT_DIR, 'main.js')
    with open(main_js_path, 'r', encoding='utf-8') as f:
        content = f.read()

    start_idx = content.find("const VLOGS = [")
    if start_idx == -1:
        print("VLOGS array niet gevonden in main.js")
        sys.exit(1)

    end_search = content[start_idx:]
    depth = 0
    end_idx = 0
    in_string = False
    escape_next = False
    for i, c in enumerate(end_search):
        if escape_next:
            escape_next = False
            continue
        if in_string:
            if c == '\\':
                escape_next = True
            elif c in '"\'':
                in_string = False
            continue
        if c in '"\'':
            in_string = True
            continue
        if c == '[':
            depth += 1
        elif c == ']':
            depth -= 1
            if depth == 0 and i + 1 < len(end_search) and end_search[i + 1] == ';':
                end_idx = start_idx + i + 2
                break

    if end_idx == 0:
        end_idx = content.find("];", start_idx) + 2

    new_content = content[:start_idx] + vlogs_block + content[end_idx:]
    with open(main_js_path, 'w', encoding='utf-8') as f:
        f.write(new_content)

    print(f"Geïmporteerd: {len(all_vlogs)} vlogs naar main.js")


if __name__ == '__main__':
    main()
